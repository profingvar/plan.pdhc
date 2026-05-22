"""Concept catalogue bulk importer (ticket #134).

Imports a list of concept rows from .xlsx or .csv. Idempotent on
concept_name: existing rows are updated (vers_number bumped); new rows
are inserted. canonical_lib + canonical_refnumber are treated as
identity fields — a row that changes either is flagged as a conflict
rather than silently overwriting.

Per the provider-integration audit (vers2 Phase 2-3), this is the
endpoint a new provider hits with a single .xlsx of concepts when
onboarding.

The importer is pure logic: it accepts a list of row dicts plus
operator/filename/sha256 metadata, and returns a JSON-shaped report.
File parsing (xlsx/csv) lives in dedicated helpers so the same code
serves the HTTP endpoint, the CLI, and tests.
"""
import csv
import hashlib
import io
import logging
import re

from app import db
from app.models.concept_models import (
    Concept, CanonicalLib, ConceptType, ResponseType, Unit,
)

logger = logging.getLogger(__name__)


REQUIRED_COLUMNS = [
    'concept_name',
    'display_text',
    'canonical_lib',
    'canonical_ref',
    'concept_type',
    'response_type',
    'unit',
]
OPTIONAL_COLUMNS = [
    'range_low',
    'range_high',
]
ALL_COLUMNS = REQUIRED_COLUMNS + OPTIONAL_COLUMNS

# concept_name must be lowercase letters, digits, and hyphens
CONCEPT_NAME_RE = re.compile(r'^[a-z0-9]+(?:-[a-z0-9]+)*$')


class ImportError_(Exception):
    """Raised when the file itself is malformed (wrong columns, etc.)."""


# ---------------------------------------------------------------------------
# File parsers
# ---------------------------------------------------------------------------

def parse_xlsx(stream):
    """Parse an .xlsx file-like into a list of row dicts.

    The first row is the header. openpyxl is imported lazily so the
    importer module loads even on environments without it (e.g.
    minimum-deps installs) — callers that hit this path will get a
    clear ImportError_.
    """
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise ImportError_(
            'openpyxl is required to import .xlsx files. '
            'Install it or upload a .csv instead.'
        )

    wb = load_workbook(stream, read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)

    try:
        header = next(rows_iter)
    except StopIteration:
        raise ImportError_('Workbook is empty')

    headers = [str(h).strip() if h is not None else '' for h in header]
    _validate_headers(headers)

    rows = []
    for raw in rows_iter:
        if raw is None or all(v is None or str(v).strip() == '' for v in raw):
            continue
        rows.append({headers[i]: raw[i] for i in range(len(headers))
                     if i < len(raw)})
    return rows


def parse_csv(stream):
    """Parse a CSV stream into a list of row dicts."""
    text = stream.read()
    if isinstance(text, bytes):
        text = text.decode('utf-8-sig')
    reader = csv.DictReader(io.StringIO(text))
    if reader.fieldnames is None:
        raise ImportError_('CSV is empty or has no header row')
    headers = [h.strip() for h in reader.fieldnames]
    _validate_headers(headers)
    rows = []
    for raw in reader:
        if all((v is None or str(v).strip() == '') for v in raw.values()):
            continue
        rows.append({k.strip(): v for k, v in raw.items()})
    return rows


def _validate_headers(headers):
    missing = [c for c in REQUIRED_COLUMNS if c not in headers]
    if missing:
        raise ImportError_(
            f'Missing required column(s): {", ".join(missing)}. '
            f'Expected: {", ".join(ALL_COLUMNS)}'
        )


# ---------------------------------------------------------------------------
# Import logic
# ---------------------------------------------------------------------------

def compute_sha256(data_bytes):
    h = hashlib.sha256()
    h.update(data_bytes)
    return h.hexdigest()


def _clean(v):
    if v is None:
        return None
    s = str(v).strip()
    return s if s != '' else None


def _to_float(v, field_name):
    v = _clean(v)
    if v is None:
        return None
    try:
        return float(v)
    except (ValueError, TypeError):
        raise ValueError(f'{field_name} is not a number: {v!r}')


def _resolve_lookup(model, name_field, value):
    """Resolve a lookup by name → row, or None if not found.

    Accepts either the name (canonical_lib_name etc.) or the GUID
    directly — operators sometimes export with GUIDs, sometimes with
    human names. Matching by GUID first is a small perf win and means
    a name that happens to look like a UUID still resolves
    correctly via the name path.
    """
    if value is None:
        return None
    value = str(value).strip()
    if value == '':
        return None
    # Try GUID match first
    row = model.query.filter_by(guid=value).first()
    if row:
        return row
    return model.query.filter(
        getattr(model, name_field) == value
    ).first()


def validate_and_import(rows, *, operator, filename, sha256, dry_run=False):
    """Validate and upsert each row. Returns a report dict.

    Args:
        rows: List of row dicts (output of parse_xlsx / parse_csv).
        operator: Identifier of the user driving the import (audit).
        filename: Original filename for audit.
        sha256: SHA-256 of the file bytes for audit.
        dry_run: If True, roll back without committing — used by the
            admin "validate-only" flow.

    Returns:
        dict: {
            'accepted': [<concept_name>, ...],
            'rejected': [{'row': <int>, 'concept_name': <str|None>,
                          'reason': <str>}, ...],
            'summary': {'n_in': N, 'n_accepted': A, 'n_rejected': R,
                        'n_created': C, 'n_updated': U,
                        'dry_run': bool},
            'operator': <str>, 'filename': <str>, 'sha256': <str>,
        }
    """
    accepted = []
    rejected = []
    n_created = 0
    n_updated = 0

    for idx, raw in enumerate(rows, start=2):  # row 1 is header
        name = _clean(raw.get('concept_name'))
        try:
            self_changed = _apply_row(raw, idx)
        except ValueError as e:
            rejected.append({'row': idx, 'concept_name': name, 'reason': str(e)})
            continue
        except _ConflictError as e:
            rejected.append({'row': idx, 'concept_name': name, 'reason': str(e)})
            continue

        accepted.append(name)
        if self_changed == 'created':
            n_created += 1
        else:
            n_updated += 1

    if dry_run:
        db.session.rollback()
    else:
        try:
            db.session.commit()
        except Exception as e:
            db.session.rollback()
            logger.exception('concept import commit failed')
            # On commit failure, treat all "accepted" rows as rejected — the
            # caller can retry. (A constraint violation here is exceptional
            # because per-row validation should have caught it.)
            return {
                'accepted': [],
                'rejected': [
                    {'row': 'commit', 'concept_name': None,
                     'reason': f'database commit failed: {e}'}
                ],
                'summary': {
                    'n_in': len(rows), 'n_accepted': 0,
                    'n_rejected': len(rows),
                    'n_created': 0, 'n_updated': 0, 'dry_run': False,
                },
                'operator': operator, 'filename': filename, 'sha256': sha256,
            }

    report = {
        'accepted': accepted,
        'rejected': rejected,
        'summary': {
            'n_in': len(rows),
            'n_accepted': len(accepted),
            'n_rejected': len(rejected),
            'n_created': n_created,
            'n_updated': n_updated,
            'dry_run': dry_run,
        },
        'operator': operator,
        'filename': filename,
        'sha256': sha256,
    }

    logger.info(
        'concept_import operator=%s filename=%s sha256=%s '
        'n_in=%d n_accepted=%d n_rejected=%d n_created=%d n_updated=%d dry_run=%s',
        operator, filename, sha256,
        report['summary']['n_in'],
        report['summary']['n_accepted'],
        report['summary']['n_rejected'],
        report['summary']['n_created'],
        report['summary']['n_updated'],
        dry_run,
    )
    return report


class _ConflictError(Exception):
    """Identity-field conflict on an existing concept."""


def _apply_row(raw, row_idx):
    """Validate + upsert a single row. Returns 'created' or 'updated'."""
    name = _clean(raw.get('concept_name'))
    if not name:
        raise ValueError('concept_name is required')
    if not CONCEPT_NAME_RE.match(name):
        raise ValueError(
            f'concept_name must be lowercase letters, digits, and hyphens '
            f'(got {name!r})'
        )

    canon_lib_val = _clean(raw.get('canonical_lib'))
    if not canon_lib_val:
        raise ValueError('canonical_lib is required')
    canon_lib_row = _resolve_lookup(CanonicalLib, 'canonical_lib_name', canon_lib_val)
    if not canon_lib_row:
        raise ValueError(f'canonical_lib not found: {canon_lib_val!r}')

    canon_ref = _clean(raw.get('canonical_ref'))
    display_text = _clean(raw.get('display_text'))

    concept_type_val = _clean(raw.get('concept_type'))
    concept_type_row = None
    if concept_type_val:
        concept_type_row = _resolve_lookup(
            ConceptType, 'concept_type_name', concept_type_val
        )
        if not concept_type_row:
            raise ValueError(f'concept_type not found: {concept_type_val!r}')

    response_type_val = _clean(raw.get('response_type'))
    response_type_row = None
    if response_type_val:
        response_type_row = _resolve_lookup(
            ResponseType, 'response_type_name', response_type_val
        )
        if not response_type_row:
            raise ValueError(f'response_type not found: {response_type_val!r}')

    unit_val = _clean(raw.get('unit'))
    unit_row = None
    if unit_val:
        unit_row = _resolve_lookup(Unit, 'unit_name', unit_val)
        if not unit_row:
            raise ValueError(f'unit not found: {unit_val!r}')

    range_low = _to_float(raw.get('range_low'), 'range_low')
    range_high = _to_float(raw.get('range_high'), 'range_high')
    if range_low is not None and range_high is not None and range_low > range_high:
        raise ValueError(f'range_low ({range_low}) > range_high ({range_high})')

    existing = Concept.query.filter_by(concept_name=name).first()
    if existing:
        # Identity-field guard: canonical_lib and canonical_refnumber are
        # append-only per the ticket. If either changes on an existing row,
        # flag a conflict so the operator decides whether to rename it or
        # update by GUID via the single-concept PUT endpoint.
        if existing.canonical_lib != canon_lib_row.guid:
            raise _ConflictError(
                f'canonical_lib conflict for {name!r}: '
                f'existing {existing.canonical_lib} vs new {canon_lib_row.guid}'
            )
        if canon_ref is not None and existing.canonical_refnumber is not None \
                and existing.canonical_refnumber != canon_ref:
            raise _ConflictError(
                f'canonical_ref conflict for {name!r}: '
                f'existing {existing.canonical_refnumber!r} vs new {canon_ref!r}'
            )
        # Update non-identity fields
        if canon_ref is not None:
            existing.canonical_refnumber = canon_ref
        if display_text is not None:
            existing.concept_display_text = display_text
        if concept_type_row:
            existing.concept_type = concept_type_row.guid
        if response_type_row:
            existing.response_type = response_type_row.guid
        if unit_row:
            existing.unit = unit_row.guid
        if range_low is not None:
            existing.range_low = range_low
        if range_high is not None:
            existing.range_high = range_high
        existing.vers_number = (existing.vers_number or 1) + 1
        return 'updated'

    # Create new
    concept = Concept(
        concept_name=name,
        canonical_lib=canon_lib_row.guid,
        canonical_refnumber=canon_ref,
        concept_display_text=display_text,
        concept_type=concept_type_row.guid if concept_type_row else None,
        response_type=response_type_row.guid if response_type_row else None,
        unit=unit_row.guid if unit_row else None,
        range_low=range_low,
        range_high=range_high,
        status='draft',
    )
    db.session.add(concept)
    return 'created'
